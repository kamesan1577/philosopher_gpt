import os
import sys
import re
from openpyxl import load_workbook, Workbook
import openai


openai.api_key = os.getenv("OPENAI_API_KEY")


def similarity_list(guideline_text, input_text):
    if len(input_text) == 0:
        raise Exception("入力された文章がありません。")
    # print(len(input_text))
    system_prompt = {
        "role": "system",
        "content": f"""
                あなたは優秀な哲学者です。ユーザーからは{len(input_text)}個の文章のリストが与えられます、それらの文章を分析して、以下に示す判定基準と価値観がどの程度合致するかを文章ごとに判定してください。
                出力は以下に示す出力例に倣い、0.0から1.0の間の数値で{len(input_text)}個出力してください。
                出力する数値の数は、入力された文章の数と同じ数にしてください。

                判定基準:
                「{guideline_text}」

                入力例:
                「["りんごは糖分が多く含まれている","バナナはおいしい","大学いくのだるい"]」
                判定基準例:
                「りんごは甘くて美味しい」
                出力例:
                「[0.9,0.8,0.1]」

                分析サービス:
                - 入力された文章を解析し、その思考やアイデアが価値観に合致した内容を含むかを確認します。
                - 入力された思考やアイデアが、どのように関連しているかを特定します。
    """,
    }

    i = 0
    output_list = []
    generated_text = ""
    print("Input text: ", input_text)
    while len(output_list) != len(input_text):
        if i > 10:
            raise Exception("OpenAI APIのリクエストが上限に達しました。")
        # OpenAI GPTでテキスト分析
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo", messages=[system_prompt, *input_texts]
        )

        generated_text = response["choices"][0]["message"]["content"]

        # 出力結果から少数を抽出し、リストに格納する
        output_list = re.findall(r"\d+\.\d+", generated_text)
        print("Temp output: ", output_list)
        i += 1

    print("Final output: ", output_list)
    return output_list


def input_text_into_cell(text):
    # Excelファイルを読み込む
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    # テキストをセルに追記する
    sheet.append([text])

    # シートを保存
    workbook.save(filename=filename)


# コマンドライン引数のパスを取得して、Excelファイルを読み込む
args = sys.argv
if len(args) == 1:
    if os.path.exists("example.xlsx"):
        raise Exception("コマンドライン引数にExcelファイルのパスを指定してください。")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["text", "similarity", "is_match", "guideline"])
        sheet["D3"].value = "threshold"
        workbook.save(filename="example.xlsx")
        raise Exception("Excelファイルを新規作成しました。")
elif len(args) != 2:
    raise Exception("コマンドライン引数にExcelファイルのパスを指定してください。")
if not os.path.exists(args[1]):
    raise Exception("指定されたExcelファイルが存在しません。")

filename = args[1]
workbook = load_workbook(filename=filename)


workbook = load_workbook(filename=filename)
sheet = workbook.active
# guideline列にある文章を取得する
if sheet["D2"].value:
    guideline_text = sheet["D2"].value
    print("guideline_text: ", guideline_text)
else:
    guideline_text = "自然の美しさや永遠のテーマを表現しており、哲学的な要素を含む。このような思考やアイデアは、宇宙や自然の神秘性についての哲学的な思考と関連している。特に、永遠や意味の追求をテーマにした哲学者や思想と関連付けられる可能性がある。"
    print("guideline_text(Default): ", guideline_text)
if sheet["D4"].value:
    if re.match(r"\d+\.\d+", str(sheet["D4"].value)):
        threshold = float(sheet["D4"].value)
        print("threshold: ", threshold)
    else:
        print(sheet["D4"].value)
        raise Exception("thresholdには0.0から1.0の間の数値を入力してください。")
else:
    threshold = 0.7
    print("threshold(Default): ", threshold)
input_texts = []
# 各セルのテキストを読み込んで分析する
for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
    if row[0] is not None:
        user_prompt = {"role": "user", "content": row[0]}
        input_texts.append(user_prompt)

sims = similarity_list(guideline_text, input_texts)

for row, sim in zip(sheet.iter_rows(min_row=2), sims):
    print(row[0].value, sim)
    row[1].value = sim
    if float(sim) >= threshold:
        row[2].value = "○"
    else:
        row[2].value = "✕"
# シートを保存
workbook.save(filename=filename)

print("Finished!")
